import openpyxl

#Reading SpreadSheet
inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

#Dictionaries
products_per_supplier = {}
total_value_per_supplier = {}
low_stock_inventory = {}

#Iteraiting through the spreadsheet
for product_n in range(2,product_list.max_row+1):

    supplier_name = product_list.cell(product_n,4).value

    # Calculating total value of inventory per supplier
    inventory = product_list.cell(product_n, 2).value
    price = product_list.cell(product_n, 3).value
    inventory_value = inventory * price

    #Calculating number of products per supplier
    if supplier_name in  products_per_supplier:
        products_per_supplier[supplier_name] += 1
    else:
        print("adding a new supplier")
        products_per_supplier[supplier_name] = 1

    # Calculating Total Value  Inventory per supplier
    if supplier_name in total_value_per_supplier:
        total_value_per_supplier[supplier_name] += inventory_value
    else:
        total_value_per_supplier[supplier_name] = inventory_value

    # Low Stock Inventory
    if inventory < 10:
        product_id = product_list.cell(product_n,1).value
        print(f" {supplier_name}: Low stock for product Id: {product_id}, available:{inventory}")
        low_stock_inventory[int(product_id)] = int(inventory)

    # Adding a New column with the total inventory value to the spreadsheet
    inventory_price_sheet = product_list.cell(product_n, 5)
    inventory_price_sheet.value = inventory_value

#Saving File in a New File
inv_file.save("Inventory_with_Total.xlsx")


print(products_per_supplier)
print(total_value_per_supplier)
print(low_stock_inventory)




